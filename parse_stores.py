from openpyxl import load_workbook


class Store:
    def __init__(self, store_number, store_format, address, city, state, zip_code, som, manager, manager_number):
        self.store_number = store_number
        self.store_format = store_format
        self.address = address
        self.city = city
        self.state = state
        self.zip_code = zip_code
        self.som = som
        self.manager = manager
        self.manager_number = manager_number

    def __str__(self):
        return "Store %s" % self.store_number

    def __repr__(self):
        return "<Store %s>" % self.store_number


def get_stores():
    wb = load_workbook('master_list.xlsx', data_only=True)
    ws = wb['master']
    total_rows = ws.max_row
    # total_rows = 10

    stores = []

    for i in range(2, total_rows):
        store_number = ws.cell(row=i, column=1).value
        store_format = ws.cell(row=i, column=11).value
        address = ws.cell(row=i, column=26).value
        city = ws.cell(row=i, column=27).value
        state = ws.cell(row=i, column=28).value
        zip_code = ws.cell(row=i, column=29).value
        som = ws.cell(row=i, column=38).value
        manager = ws.cell(row=i, column=36).value
        manager_number = ws.cell(row=i, column=37).value

        stores.append(
            Store(
                store_number=store_number,
                store_format=store_format,
                address=address,
                city=city,
                state=state,
                zip_code=zip_code,
                som=som,
                manager=manager,
                manager_number=manager_number
            )
        )

    return stores


print(get_stores()[0].address)
